import openpyxl, datetime

wb = openpyxl.load_workbook('2022-23 5th BDS Timetable V06.xlsx')

def GetAllDates():
    for sheet in wb.worksheets:
        print(sheet.title)
        if "wc" in sheet.title:
            pass

def range_char(start, stop):
    return (chr(n) for n in range(ord(start), ord(stop) + 1))

def GetWeek(RowNumber, sheet):
    Week_List = []
    Day_List = []
    EventDay = sheet['E3'].value

    for i in range_char("e", "s"):
        MergedCell = False
        for rng in sheet.merged_cells.ranges:
            if i + str(RowNumber) in rng:
                MergedCell = True
                match len(Day_List):
                    case 0:
                        EventDay = EventDay.replace(hour = 9, minute = 0)
                        EventEnd = EventDay.replace(hour = 12, minute = 30)
                    case 1:
                        if (sheet[i + str(RowNumber)].value == 'IOC'):
                            EventDay = EventDay.replace(hour = 12, minute = 30)
                            EventEnd = EventDay.replace(hour = 17, minute = 0)
                        else:
                            EventDay = EventDay.replace(hour = 12, minute = 45)
                            EventEnd = EventDay.replace(hour = 13, minute = 45)
                    case 2:
                        if (sheet[i + str(RowNumber)].value == 'IOC'):
                            EventDay = EventDay.replace(hour = 13, minute = 45)
                            EventEnd = EventDay.replace(hour = 17, minute = 0)
                        else:
                            EventDay = EventDay.replace(hour = 13, minute = 30)
                            EventEnd = EventDay.replace(hour = 17, minute = 0)

                Event = {'Date':EventDay.date(), 'StartTime':EventDay.isoformat(), 'EndTime':EventEnd.isoformat(), 'Event':sheet[str(rng).split(':')[0]].value} #Splits value of merged cell range before ':', giving the first cell location - which it then finds the value of
                Day_List.append(Event)

        if MergedCell == False:
            match len(Day_List):
                case 0:
                    EventDay = EventDay.replace(hour = 9, minute = 0)
                    EventEnd = EventDay.replace(hour = 12, minute = 30)
                case 1:
                    if (sheet[i + str(RowNumber)].value == 'IOC'):
                        EventDay = EventDay.replace(hour = 12, minute = 30)
                        EventEnd = EventDay.replace(hour = 17, minute = 0)
                    else:
                        EventDay = EventDay.replace(hour = 12, minute = 45)
                        EventEnd = EventDay.replace(hour = 13, minute = 45)
                case 2:
                        if (sheet[i + str(RowNumber)].value == 'IOC'):
                            EventDay = EventDay.replace(hour = 13, minute = 45)
                            EventEnd = EventDay.replace(hour = 17, minute = 0)
                        else:
                            EventDay = EventDay.replace(hour = 13, minute = 30)
                            EventEnd = EventDay.replace(hour = 17, minute = 0)

            Event = {'Date':EventDay.date(), 'StartTime':EventDay.isoformat(), 'EndTime':EventEnd.isoformat(), 'Event':sheet[i + str(RowNumber)].value}
            Day_List.append(Event)

        if (len(Day_List)/3).is_integer() and len(Day_List) != 0:
            Week_List.append(Day_List)
            Day_List = []
            EventDay = EventDay + datetime.timedelta(days= 1)

    return Week_List

def GetWeekByStudentNumber(studentnumber):
    ScheduleList = []
    for sheet in wb.worksheets:
        print(sheet.title)
        if "wc" in sheet.title:
            for row in range(1, sheet.max_row):
                if sheet.cell(row=row, column = 4).value == studentnumber: #Student No place holder
                    ScheduleList = ScheduleList + GetWeek(row, sheet)

    return ScheduleList
